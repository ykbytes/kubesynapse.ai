"""MCP Documents sidecar — read and create PDF, Excel, and Word files."""

import io
import logging
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "base"))
from mcp_base import create_mcp_server, run_server

log = logging.getLogger("mcp-documents")

server = create_mcp_server(
    "mcp-documents",
    "Read and create PDF, Excel (.xlsx), and Word (.docx) documents.",
)

WORK_DIR = os.environ.get("MCP_WORK_DIR", tempfile.gettempdir())
MAX_TEXT_CHARS = 16000

# --- Path traversal protection ---
# Only files within these directories may be read or created.
_ALLOWED_BASE_DIRS = [
    Path(WORK_DIR).resolve(),
    Path("/tmp").resolve(),
]


def _validate_file_path(file_path: str) -> str | None:
    """Return an error message if file_path is outside allowed directories, else None."""
    try:
        resolved = Path(file_path).resolve()
    except (OSError, ValueError):
        return "Invalid file path"

    for base_dir in _ALLOWED_BASE_DIRS:
        try:
            resolved.relative_to(base_dir)
            if not resolved.is_file():
                return f"Path is not a file: {file_path}"
            return None
        except ValueError:
            continue

    return (
        f"File path is outside allowed directories. "
        f"Allowed: {', '.join(str(d) for d in _ALLOWED_BASE_DIRS)}"
    )


@server.tool()
def read_pdf(file_path: str) -> str:
    """Extract text content from a PDF file."""
    path_err = _validate_file_path(file_path)
    if path_err:
        return f"BLOCKED: {path_err}"
    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"--- Page {i + 1} ---\n{text}")
        return "\n".join(pages)[:MAX_TEXT_CHARS] if pages else "(no text found in PDF)"
    except ImportError:
        return "ERROR: pypdf not installed"
    except Exception as e:
        return f"ERROR: Failed to read PDF: {e}"


@server.tool()
def read_xlsx(file_path: str, sheet_name: str = "") -> str:
    """Read an Excel file and return contents as text. Optionally specify sheet_name."""
    path_err = _validate_file_path(file_path)
    if path_err:
        return f"BLOCKED: {path_err}"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
        lines = []
        for sn in sheets:
            ws = wb[sn]
            lines.append(f"=== Sheet: {sn} ===")
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(lines)[:MAX_TEXT_CHARS]
    except ImportError:
        return "ERROR: openpyxl not installed"
    except Exception as e:
        return f"ERROR: Failed to read Excel file: {e}"


@server.tool()
def read_docx(file_path: str) -> str:
    """Extract text content from a Word (.docx) file."""
    path_err = _validate_file_path(file_path)
    if path_err:
        return f"BLOCKED: {path_err}"
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)[:MAX_TEXT_CHARS] if paragraphs else "(no text found)"
    except ImportError:
        return "ERROR: python-docx not installed"
    except Exception as e:
        return f"ERROR: Failed to read docx: {e}"


@server.tool()
def create_pdf(text: str, output_filename: str = "output.pdf") -> str:
    """Create a simple PDF from text content. Returns the file path."""
    # Prevent path traversal via filename
    safe_filename = os.path.basename(output_filename)
    if not safe_filename or safe_filename != output_filename:
        return "BLOCKED: Invalid output filename (path traversal detected)"
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        path = os.path.join(WORK_DIR, safe_filename)
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4
        y = height - 50
        for line in text.split("\n"):
            if y < 50:
                c.showPage()
                y = height - 50
            c.drawString(50, y, line[:100])
            y -= 14
        c.save()
        return f"PDF created: {path}"
    except ImportError:
        return "ERROR: reportlab not installed"
    except Exception as e:
        return f"ERROR: Failed to create PDF: {e}"


if __name__ == "__main__":
    run_server(server)
